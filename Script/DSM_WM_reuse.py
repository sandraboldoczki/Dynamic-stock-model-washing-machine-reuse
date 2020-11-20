#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 19 17:11:11 2019

@author: sandraboldoczki
"""
# =============================================================================
# Dynamic Stock Model for Reuse of Waste Electrical and Electronic Equipment
# DSM 1.0 
# =============================================================================

#%%

import numpy as np
import scipy.stats
from datetime import datetime
from pathlib import Path
import os
import shutil 
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl as xl 

#%%
"""Set Paths"""
#Initialize model run and data copy
#Specify path to dynamic stock model and to datafile, relative

Path_Data = os.path.join(os.getcwd(),'..','Data')
FileName_Script = 'DSM_WM_reuse.py'
Path_Results    = Path().resolve().parents[0] / 'Results'
Path.mkdir(Path_Results, exist_ok = True)
Path_Script     = Path() / FileName_Script

##save copy of code
dtime       = datetime.now()
dtimestr    = dtime.strftime("%Y_%m_%d__%H_%M_%S")
Path_Run = Path_Results / dtimestr
Path(Path_Run).mkdir(parents=True, exist_ok=True)## subfolder created
shutil.copy(Path_Script,Path_Run) ## source + target

##Read Input Data
FileName_Data = 'Input_DSMWM.xlsx'
Path_WB     = os.path.join(Path_Data, FileName_Data)
Pro_DataFileWB = xlrd.open_workbook (Path_WB)

print("Input data set up.")

#%%
"""Define System Variables"""

Pro_DefSheet = Pro_DataFileWB.sheet_by_name('Definitions')

#If change of year: Change demand in input file!
Par_StartYear     = 1991
Par_EndYear       = 2051

Par_NoOfYears     = Par_EndYear - Par_StartYear + 1
Par_NoOfScenarios = 3

Par_CurrentYear   = 2019 

Def_Time = []
for t in range (Par_NoOfYears):
    Def_Time.append(np.int(Par_StartYear+t))
    
Par_NoOfEEC       = 10

Def_EECName       = []
Def_EECCode       = []


for e in range (Par_NoOfEEC):
    Def_EECName.append(Pro_DefSheet.cell_value(e+2,2)) ##cell_value(row_num,col_num), starting with 0
    Def_EECCode.append(Pro_DefSheet.cell_value(e+2,3))

Def_Processes = ["Production", "Use phase", "End-of-life", "Preparation for Reuse", "Recycling", "Stock Reuse", "Stock New"]       
    
print ("All system variables defined.")

#%%

"""Define model parameters"""

Pro_ParSheet = Pro_DataFileWB.sheet_by_name('Parameters')

e = 0

Par_Lifetime_New = Pro_ParSheet.cell_value(e+2,1)
Par_Lifetime_Reuse = Pro_ParSheet.cell_value(e+2,2)
Par_LTNew_Shape  = Pro_ParSheet.cell_value(e+2,3)
Par_LTReuse_Shape  = Pro_ParSheet.cell_value(e+2,4)

##Read demand parameters
Pro_DemSheet = Pro_DataFileWB.sheet_by_name('Demand')

Par_DemandbyYear = np.zeros(Par_NoOfYears)
Par_DemandbyPercent = np.zeros ((Par_NoOfYears, Par_NoOfEEC))

for t in range (0, Par_NoOfYears):
   Par_DemandbyYear [t] = Pro_DemSheet.cell_value(t+1,1)
   for e in range (0, Par_NoOfEEC):
       Par_DemandbyPercent [t,e] = Pro_DemSheet.cell_value(t+1,e+2)
   
##Determine initial stock
Pro_StockSheet = Pro_DataFileWB.sheet_by_name('InitialStock')
Par_InitialStock = np.zeros(Par_NoOfEEC)

for s in range (0, Par_NoOfEEC):
    Par_InitialStock[s] = Pro_StockSheet.cell_value (s+1,1)
    
##Define lifetime distribution
Par_PDF = np.zeros ((Par_NoOfYears,Par_NoOfYears))
Par_PDF_RU = np.zeros ((Par_NoOfYears, Par_NoOfYears))
Par_PDFArray = np.zeros ((Par_NoOfYears,Par_NoOfYears, Par_NoOfEEC))
Par_PDFArray_RU = np.zeros ((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))

for m in range (0, Par_NoOfYears):

  Par_PDF [m+1::, m] = scipy.stats.weibull_min(Par_LTNew_Shape, 0, Par_Lifetime_New).cdf(np.arange(1, Par_NoOfYears - m)) - scipy.stats.weibull_min( Par_LTNew_Shape,0,Par_Lifetime_New).cdf(np.arange(0, Par_NoOfYears-m-1))
  Par_PDF_RU [m+1::, m] = scipy.stats.weibull_min(Par_LTReuse_Shape, 0, Par_Lifetime_Reuse).cdf(np.arange(1, Par_NoOfYears - m)) - scipy.stats.weibull_min( Par_LTReuse_Shape,0,Par_Lifetime_Reuse).cdf(np.arange(0, Par_NoOfYears-m-1))

for m in range (0, Par_NoOfYears):
    for n in range (0, Par_NoOfYears):
        for e in range (0, Par_NoOfEEC):
           
            Par_PDFArray [n,m,e] = Par_PDF [n,m]
            Par_PDFArray_RU [n,m,e] = Par_PDF_RU [n,m]
            
"""Read Scenarios"""

Pro_ReuseRate = Pro_DataFileWB.sheet_by_name('Reuse')
Par_Alpha = np.zeros((Par_NoOfYears,Par_NoOfScenarios))


for a in range (0, Par_NoOfYears):
    for s in range (0, Par_NoOfScenarios):
        Par_Alpha [a,s] = Pro_ReuseRate.cell_value(a+2,s+1)

print ("Scenarios implemented.")
print ("All parameters set up.")

#%%
for sc in range (0,Par_NoOfScenarios):

    """Define System Variables"""
    
    ##Define Flows 
    
    ##Primary Material Extraction (not relevant)
    Fl_01 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    
    ##Production
    Fl_12 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    
    ##EoL
    Fl_23 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    fR_23 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    
    ##Preparation for Reuse
    Fl_34 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    
    ##Recycling 
    
    Fl_35 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    fR_35 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    
    ##ReUseCase
    
    Fl_42 = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    
    ##Sort Reuse 
    
    EoL_Buffer = np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    Par_RestDemand = np.zeros((Par_NoOfYears))
    Par_ToBeProduced = np.zeros((Par_NoOfYears))
    Stock_Difference = np.zeros((Par_NoOfYears))
    MaxReuse = np.zeros((Par_NoOfYears,Par_NoOfYears))
    ToSort = np.zeros ((Par_NoOfYears,Par_NoOfYears)) 
    Reused = np.zeros((Par_NoOfYears,Par_NoOfYears, Par_NoOfEEC))
    Par_CheckAlpha = np.zeros(Par_NoOfYears)
    
    
    print("All Flow variables defined.")
    
    ##Inflow to Stock
    
    Stock= np.zeros((Par_NoOfYears, Par_NoOfYears, Par_NoOfEEC))
    Stock_Reuse = np.zeros((Par_NoOfYears, Par_NoOfYears,Par_NoOfEEC))
    
    Total_Stock = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    Total_Stock_New = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    Total_Stock_Reuse = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    
    Total_Recycling = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    Total_Reuse = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    Total_Production = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    Total_EoL = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    OnlyNew_EoL = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    OnlyNew_Rec = np.zeros ((Par_NoOfYears, Par_NoOfEEC))
    
    print("All Stocks defined.")

#%%
    print ("Calculating Stocks..."+ str(sc))
    """Calculations"""

    for t in range (0,Par_NoOfEEC):
       Fl_12[0,0,t] = Par_InitialStock [t]
       Stock [0,0,t] = Fl_12 [0,0,t]   
       
     
    for y in range (0, Par_NoOfYears):#For all years
        for c in range (0, y+1): #Cohort c from first to last year       
                 for e in range (0, Par_NoOfEEC):
                    
              
                    
                    if c != 0: # year 0 determined prior to loop
                        Par_ToBeProduced [y] = max(Par_DemandbyYear [y] - Stock [y,:c,:].sum() - Stock_Reuse [y,:c,:].sum() - Fl_42 [y-1,:,:].sum(),0) #Calculates products to be produced
                        Fl_12[y,y,e] = Par_ToBeProduced[y] * Par_DemandbyPercent [y,e]
                        Stock[y,y,e] += Fl_12[y,y,e]
                    
                    Stock [y,c,e] = Fl_12[c,c,e] * (1 - Par_PDFArray [0:y+1,c,e].sum()) #Stock determined by Inflow
                    Stock_Reuse [y,c,e] = Fl_42[c-1,:,e].sum() * (1 - Par_PDFArray_RU [0:y+1,c,e].sum()) #Stock Reuse
                    Total_Stock [y,e] = Stock [y,:,e].sum() + Stock_Reuse [y,:,e].sum() 
                    
                    Fl_23[y,c,e] = Par_PDFArray[y,c,e] * Fl_12[c,c,e]  
                    fR_23[y,c,e] = Par_PDFArray_RU[y,c,e] * Fl_42 [c-1,:,e].sum()
                 
    
                    """Flow Recycling, Flow VzW"""
                 
                    fR_35 [y,c,e] = fR_23 [y,c,e] #Flow to Recycling
                  
                    EoL_Buffer[y,y,e] = Fl_23 [y,:,e].sum() #All products that reach their EoL go to "buffer"
                    Stock_Difference [y] = max((Par_DemandbyYear [y] - Stock [y,:c,:].sum() - Stock_Reuse [y,:c,:].sum() ),0) #Gap in in-use stock that needs to be filled 
                    Par_RestDemand [y] = max(Stock_Difference[y], 0) #defines residual demand
                    
                    MaxReuse [y,y] = Par_Alpha [y,sc] * Fl_23[y,:,:].sum() #maximum reuse quota, dependent on alpha
                    
                    Par_CheckAlpha [y] = Par_Alpha[y,sc]
                    ToSort [y,y] = MaxReuse [y,y] #Products that need to be sorted
                    
                    if Par_RestDemand[y]< MaxReuse [y,y]: #If residual demand smaller than maximum reuse quota,then only reuse demanded products, rest goes to recyclingn
                        
                        if e == 0:
                            print ("Oversaturation in Demand in Year "+str(Par_CurrentYear+y))
                        ToSort[y,y]=Par_RestDemand[y]
     
                   
                    for n in range (0,y):
                        for s in range (0, Par_NoOfEEC):
                            
                       
                                if EoL_Buffer [n,n,s] > 0:
                                        Reused [n,n,s] =  min (ToSort[n,n], EoL_Buffer[n,n,s]) #Minimum of products to be sorted and all EoL products
                                        Fl_34 [n,n,s] += Reused [n,n,s] 
                                        Fl_42 [n,n,s] = Fl_34 [n,n,s] 
                                        EoL_Buffer [n,n,s] = EoL_Buffer [n,n,s] - Reused [n,n,s] #Reused products are substracted from "buffer"
                                        Fl_35 [n,n,s] = EoL_Buffer [n,n,s] #Rest goes to recyling
                                        EoL_Buffer [n,n,s] -= Fl_35 [n,n,s] #"Buffer" should be zero afterwards
                                        ToSort [n,n] -=  Reused [n,n,s] 
                                       
                       
                    Total_Stock_New [y,e] = Stock [y,:,e].sum() #Sum over all cohorts per year
                    Total_Stock_Reuse [y,e] = Stock_Reuse [y,:,e].sum()
                    
                    Total_Reuse [c,e] = Fl_42[c,:,e].sum()
                    Total_Recycling [c,e] = Fl_35[c,:,e].sum()+ fR_35[c,:,e].sum()
                    Total_Production[c,e] = Fl_12[c,:,e].sum() 
                    Total_EoL[c,e] = Fl_23 [c,:,e].sum()+ fR_23[c,:,e].sum() #Total EoL also contains reused products
                    OnlyNew_EoL[c,e] = Fl_23 [c,:,e].sum() 
                    OnlyNew_Rec[c,e] = Fl_35[c,:,e].sum()
                 
    print("Finished calculations.")      
    
    
    #%%
    
    """Define model parameters LCA"""
    
    Pro_DefSheetLCA = Pro_DataFileWB.sheet_by_name('DefLCA')
    
    Par_NoOfImpactCat     = 25
    Def_ImpactCatName     = []
    Def_ImpactCatCode     = []
    
    
    for f in range (Par_NoOfImpactCat):
        Def_ImpactCatName.append(Pro_DefSheetLCA.cell_value(f+2,2)) #cell_value(row_num,col_num), starting with 0
        Def_ImpactCatCode.append(Pro_DefSheetLCA.cell_value(f+2,3))
    
    Pro_DemandWE = Pro_DataFileWB.sheet_by_name('DemandWE')
     
    Def_DemandWater          = []
    Def_DemandElectricity    = []
    Def_DemandProduction    = []
    Def_DemandRecycling    = []
    
    for e in range (Par_NoOfEEC):
        Def_DemandWater.append(Pro_DemandWE.cell_value(1,e+2)) #cell_value(row_num,col_num), starting with 0
        Def_DemandElectricity.append(Pro_DemandWE.cell_value(2,e+2))
        Def_DemandProduction.append(Pro_DemandWE.cell_value(7,e+2))
        Def_DemandRecycling.append(Pro_DemandWE.cell_value(8,e+2)) 
        
    Def_UB = 220 #Usage behavior, e.g. 220 washing cycles per year
    
    ##Read impact parameters per product
        
    ##Impacts of water
    Pro_IWSheet = Pro_DataFileWB.sheet_by_name('ImpactsWater')
    Par_ImpactsWater = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    
    for t in range (Par_NoOfYears-24):
       Par_ImpactsWater [24+t] = Pro_IWSheet.cell_value(t+2,1)
       for e in range (0, Par_NoOfImpactCat):
           Par_ImpactsWater [24+t,e] = Pro_IWSheet.cell_value(t+2,e+1)
    
    ##Impacts of electricity
    Pro_IESheet = Pro_DataFileWB.sheet_by_name('ImpactsElectricity')
    Par_ImpactsElectricity = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    
    for t in range (Par_NoOfYears-24):
       Par_ImpactsElectricity [24+t] = Pro_IESheet.cell_value(t+2,1)
       for e in range (0, Par_NoOfImpactCat):
           Par_ImpactsElectricity [24+t,e] = Pro_IESheet.cell_value(t+2,e+1)
    
    ##Impacts per process for single products
    Pro_IProcSheet = Pro_DataFileWB.sheet_by_name('ImpactsProcesses')
    
    ITotalnew = np.zeros(Par_NoOfImpactCat) #Total new per product
    IP = np.zeros(Par_NoOfImpactCat) #Production per product
    ITn =  np.zeros(Par_NoOfImpactCat) #Transport new per product
    ITotalPfR = np.zeros(Par_NoOfImpactCat) #Total PfR per product
    IR =  np.zeros(Par_NoOfImpactCat) #Repair per product
    ITPfR =  np.zeros(Par_NoOfImpactCat) #Transport PfR per product
    IRec =  np.zeros(Par_NoOfImpactCat) #Recycling per product
    
    for f in range (Par_NoOfImpactCat):
        ITotalnew[f] = np.asarray(Pro_IProcSheet.cell_value(2,2+f)) #Total new per product
        IP[f] = np.asarray(Pro_IProcSheet.cell_value(3,2+f)) #Production per product
        ITn[f] = np.asarray(Pro_IProcSheet.cell_value(4,2+f)) #Transport new per product
        ITotalPfR[f] = np.asarray(Pro_IProcSheet.cell_value(5,2+f)) #Total PfR per product
        IR[f] = np.asarray(Pro_IProcSheet.cell_value(6,2+f)) #Repair per product
        ITPfR[f] = np.asarray(Pro_IProcSheet.cell_value(7,2+f)) #Transport PfR per product
        IRec[f] = np.asarray(Pro_IProcSheet.cell_value(8,2+f)) #Recycling per product
        
        
    """Define System Variables for LCA"""
    
    ##Define Impacts per process and year for all products
    
    Imp_Total = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_Totalnew = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_P = np.zeros((Par_NoOfYears, Par_NoOfImpactCat, Par_NoOfEEC))
    Imp_Pperyear = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_Tnew = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_U = np.zeros((Par_NoOfYears, Par_NoOfImpactCat, Par_NoOfEEC))
    Imp_TotalPfR = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_R = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_TPfR = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_Rec = np.zeros((Par_NoOfYears, Par_NoOfImpactCat,Par_NoOfEEC))
    Imp_Recperyear = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    Imp_allperYear = np.zeros((Par_NoOfYears, Par_NoOfImpactCat))
    
    print("All Impacts defined.")
    
    #%%
    print ("Calculating Impacts...")
    """Calculations"""
    
    for t in range (24, Par_NoOfYears):
       for e in range (0, Par_NoOfImpactCat):
          
          
           Imp_Tnew[t,e] = ITn[e]*Total_Production.sum(axis=1)[t]
           Imp_R[t,e] = IR[e]*Total_Reuse.sum(axis=1)[t]
           Imp_TPfR[t,e] = ITPfR[e]*Total_Reuse.sum(axis=1)[t]

        
        
    for t in range (24, Par_NoOfYears):
       for e in range (0, Par_NoOfImpactCat):
          for s in range (0, Par_NoOfEEC):  
         
           Imp_P[t,e,s] = IP[e]*Def_DemandProduction[s]*Total_Production[t,s]  #impacts production per EEC   
           Imp_Pperyear[t,e] = Imp_P.sum(axis=2)[t,e]
           Imp_Totalnew [t,e] = Imp_Tnew[t,e] +  Imp_Pperyear[t,e]
           Imp_Rec[t,e,s] = IRec[e]*Def_DemandRecycling[s]*Total_Recycling[t,s]
           Imp_Recperyear[t,e] = Imp_Rec.sum(axis=2)[t,e]
           Imp_TotalPfR [t,e] = Imp_R[t,e] + Imp_TPfR[t,e]
           Imp_U[t,e,s] = ((Par_ImpactsWater[t,e]*Def_DemandWater[s]+ Par_ImpactsElectricity[t,e]*Def_DemandElectricity[s])*Def_UB*Total_Stock[t,s]) #impacts use per EEC
    

    Def_YIU = [] #yearly Impacts Use
    Def_YIU = np.sum(a =Imp_U, axis = 2 )
    
    
    for t in range (24, Par_NoOfYears):
       for e in range (0, Par_NoOfImpactCat):
          Imp_allperYear [t,e] = Imp_Totalnew[t,e] + Imp_TotalPfR[t,e]
          Imp_Total[t,e] = Def_YIU[t,e] + Imp_Totalnew[t,e] + Imp_TotalPfR[t,e] +  Imp_Recperyear[t,e]
    
    
    print("Finished calculations.")      
    
    #%%

    
    Result_WB = Workbook()
    
    ws1 = Result_WB.active
    ws1.title = "Resulting Stocks"
    
    myfont = Font(name='Arial',
                     size=13,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000') 
    
    bd = Border(left=Side(style=None), 
                         right=Side(style='thick'), 
                         top=Side(style=None), 
                         bottom=Side(style=None))
    
    ws1['A1'] = "Dynamic Stock Model for WEEE"
    ws1['A2'] = "Resulting Stocks"
    ws1['E1'] = "Created: "+ str(datetime.now())
    ws1['E2'] = "Selected Reuse Scenario: Scenario "+str(sc+1)
    
    ws1['A1'].font = myfont
    ws1['A2'].font = myfont
    ws1['E1'].font = myfont
    #Jahre
    for a in range (24, Par_NoOfYears-1): 
        ws1.cell(column=1, row = a+6-24, value = Def_Time[a])
        ws1.cell(column=1, row = a+6-24).font = myfont 
    
        
    for p in range(len(Def_Processes)):
        for e in range(0, Par_NoOfEEC):
            ws1.cell(column = 2+e*7, row=4, value = Def_EECName[e])
            ws1.cell(column = 2+p+e*7, row=5, value = Def_Processes[p])
            
            ws1.cell(column = 8+e*7, row=4).border = bd
            ws1.cell(column = 8+e*7, row=5).border = bd
            
            ws1.cell(column = 2+e*7, row=4).font = myfont 
            ws1.cell(column = 2+e*7, row=5).font = myfont
            ws1.cell(column = 3+e*7, row=5).font = myfont
            ws1.cell(column = 4+e*7, row=5).font = myfont
            ws1.cell(column = 5+e*7, row=5).font = myfont
            ws1.cell(column = 6+e*7, row=5).font = myfont
            ws1.cell(column = 7+e*7, row=5).font = myfont
            ws1.cell(column = 8+e*7, row=5).font = myfont
            
            ws1.cell(column = 72, row=5).font = myfont
            ws1.cell(column = 73, row=5).font = myfont
            ws1.cell(column = 74, row=5).font = myfont
            ws1.cell(column = 72, row=5, value = "Overall Production")
            ws1.cell(column = 73, row=5, value = "Overall Reuse")
            ws1.cell(column = 74, row=5, value = "Overall Recycling")
    
    
    
    for t in range (24, Par_NoOfYears-1):
        for e in range (0, Par_NoOfEEC):
            ws1.cell(column = 2+e*7, row=6+t-24, value = Total_Production[t,e])
            ws1.cell(column = 3+e*7, row=6+t-24, value = Total_Stock[t,e])
            ws1.cell(column = 4+e*7, row=6+t-24, value = Total_EoL[t,e])
            ws1.cell(column = 5+e*7, row=6+t-24, value = Total_Reuse[t,e])
            ws1.cell(column = 6+e*7, row=6+t-24, value = Total_Recycling[t,e])
            ws1.cell(column = 7+e*7, row=6+t-24, value = Total_Stock_Reuse[t,e])
            ws1.cell(column = 8+e*7, row=6+t-24, value = Total_Stock_New[t,e])
            ws1.cell(column = 8+e*7, row=6+t-24).border = bd
            
            
    for t in range (24, Par_NoOfYears-1):
        ws1.cell(column = 72, row = 6+t-24, value = Total_Production[t,:].sum())
        ws1.cell(column = 73, row = 6+t-24, value = Total_Reuse[t,:].sum())
        ws1.cell(column = 74, row = 6+t-24, value = Total_Recycling[t,:].sum())
        
    #Generate new sheet with data for sankey
    
    ws2 =  Result_WB.create_sheet('sheet 2')
    ws2.title = "Data Sankey"
    
    myfont = Font(name='Arial',
                     size=13,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000') 
    
    bd = Border(left=Side(style=None), 
                         right=Side(style='thick'), 
                         top=Side(style=None), 
                         bottom=Side(style=None))
    ws2['A1'] = "Dynamic Stock Model for WEEE"
    ws2['A2'] = "Data for Sankey"
    ws2['E1'] = "Created: "+ str(datetime.now())
    ws2['E2'] = "Selected Reuse Scenario: Scenario "+str(sc+1)
    
    ws2['A1'].font = myfont
    ws2['A2'].font = myfont
    ws2['E1'].font = myfont
    
    ##ONLY NEW by years
    for a in range (24, Par_NoOfYears-1): 
         for d in range (0, 5):
             ws2.cell(column=1+d*11, row = a+6-24, value = Def_Time[a])
             ws2.cell(column=1+d*11, row = a+6-24).font = myfont 
    
    #Name Process und Energy efficiency class
    #for e in range (0, Par_NoOfEEC):
    ws2.cell(column = 2+0*11, row=4, value = "Total_Stock_New")
    ws2.cell(column = 2+1*11, row=4, value = "OnlyNew_EoL")
    ws2.cell(column = 2+2*11, row=4, value = "OnlyNew_Recycling")
    ws2.cell(column = 2+3*11, row=4, value = "OnlyNew_PfR")
    ws2.cell(column = 2+4*11, row=4, value = "Production_New")
    
    ws2.cell(column = 2+0*11, row=4).font = myfont
    ws2.cell(column = 2+1*11, row=4).font = myfont 
    ws2.cell(column = 2+2*11, row=4).font = myfont 
    ws2.cell(column = 2+3*11, row=4).font = myfont 
    ws2.cell(column = 2+4*11, row=4).font = myfont  
    
    for e in range (0, Par_NoOfEEC):
        ws2.cell(column = 2+0*11+e, row=5, value = Def_EECName[e])
        ws2.cell(column = 2+1*11+e, row=5, value = Def_EECName[e])
        ws2.cell(column = 2+2*11+e, row=5, value = Def_EECName[e])  
        ws2.cell(column = 2+3*11+e, row=5, value = Def_EECName[e])
        ws2.cell(column = 2+4*11+e, row=5, value = Def_EECName[e])
        
    #Data
    for t in range (24, Par_NoOfYears-1):
     for e in range (0, Par_NoOfEEC):
         
        ws2.cell(column = 2+0*11+e, row=6+t-24, value = Total_Stock_New[t,e])
        ws2.cell(column = 2+1*11+e, row=6+t-24, value = OnlyNew_EoL[t,e])
        ws2.cell(column = 2+2*11+e, row=6+t-24, value = OnlyNew_Rec[t,e])
        ws2.cell(column = 2+3*11+e, row=6+t-24, value = OnlyNew_EoL[t,e]-OnlyNew_Rec[t,e])
        ws2.cell(column = 2+4*11+e, row=6+t-24, value = Total_Production[t,e])
        
        ws2.cell(column = 11+0*11, row=6+t-24).border = bd
        ws2.cell(column = 11+1*11, row=6+t-24).border = bd
        ws2.cell(column = 11+2*11, row=6+t-24).border = bd
        ws2.cell(column = 11+3*11, row=6+t-24).border = bd
        ws2.cell(column = 11+4*11, row=6+t-24).border = bd
    ##End ONLY NEW
    
    ##ONLY PfR by years
    for a in range (24, Par_NoOfYears-1): 
         for d in range (0, 5):
             ws2.cell(column=1+d*11, row = a+46-24, value = Def_Time[a])
             ws2.cell(column=1+d*11, row = a+46-24).font = myfont 
    
    #Name Process und Energy efficiency class
    ws2.cell(column = 2+0*11, row=44, value = "Total_Stock_Reuse")
    ws2.cell(column = 2+1*11, row=44, value = "OnlyReuse_EoL")
    ws2.cell(column = 2+2*11, row=44, value = "OnlyReuse_Recycling")
    ws2.cell(column = 2+3*11, row=44, value = "OnlyReuse_PfR")
    ws2.cell(column = 2+4*11, row=44, value = "Production_PfR")
    
    ws2.cell(column = 2+0*11, row=44).font = myfont
    ws2.cell(column = 2+1*11, row=44).font = myfont 
    ws2.cell(column = 2+2*11, row=44).font = myfont 
    ws2.cell(column = 2+3*11, row=44).font = myfont 
    ws2.cell(column = 2+4*11, row=44).font = myfont  
    
    for e in range (0, Par_NoOfEEC):
        ws2.cell(column = 2+0*11+e, row=45, value = Def_EECName[e])
        ws2.cell(column = 2+1*11+e, row=45, value = Def_EECName[e])
        ws2.cell(column = 2+2*11+e, row=45, value = Def_EECName[e])  
        ws2.cell(column = 2+3*11+e, row=45, value = Def_EECName[e])
        ws2.cell(column = 2+4*11+e, row=45, value = Def_EECName[e])
        
    #Data
    for t in range (24, Par_NoOfYears-1):
     for e in range (0, Par_NoOfEEC):
         
        ws2.cell(column = 2+0*11+e, row=46+t-24, value = Total_Stock_Reuse[t,e])
        ws2.cell(column = 2+1*11+e, row=46+t-24, value = Total_EoL[t,e]-OnlyNew_EoL[t,e])
        ws2.cell(column = 2+2*11+e, row=46+t-24, value = Total_Recycling[t,e]-OnlyNew_Rec[t,e])
        ws2.cell(column = 2+3*11+e, row=46+t-24, value = Total_Reuse[t,e]-OnlyNew_EoL[t,e]-OnlyNew_Rec[t,e])
        ws2.cell(column = 2+4*11+e, row=46+t-24, value = Total_Production[t,e]-Total_Production[t,e])
        
        ws2.cell(column = 11+0*11, row=46+t-24).border = bd
        
        ws2.cell(column = 11+1*11, row=46+t-24).border = bd
        ws2.cell(column = 11+2*11, row=46+t-24).border = bd
        ws2.cell(column = 11+3*11, row=46+t-24).border = bd
        ws2.cell(column = 11+4*11, row=46+t-24).border = bd   
    
    ##End ONLY PfR
    
    ##Total by Years
    for a in range (24, Par_NoOfYears-1): 
         for d in range (0, 5):
             ws2.cell(column=1+d*11, row = a+86-24, value = Def_Time[a])
             ws2.cell(column=1+d*11, row = a+86-24).font = myfont 
    
    #Name Process und Energy efficiency class
    ws2.cell(column = 2+0*11, row=84, value = "Total_Stock")
    ws2.cell(column = 2+1*11, row=84, value = "Total_EoL")
    ws2.cell(column = 2+2*11, row=84, value = "Total_Recycling")
    ws2.cell(column = 2+3*11, row=84, value = "Total_PfR")
    ws2.cell(column = 2+4*11, row=84, value = "Total_Production")
    
    ws2.cell(column = 2+0*11, row=84).font = myfont
    ws2.cell(column = 2+1*11, row=84).font = myfont 
    ws2.cell(column = 2+2*11, row=84).font = myfont 
    ws2.cell(column = 2+3*11, row=84).font = myfont 
    ws2.cell(column = 2+4*11, row=84).font = myfont  
    
    for e in range (0, Par_NoOfEEC):
        ws2.cell(column = 2+0*11+e, row=85, value = Def_EECName[e])
        ws2.cell(column = 2+1*11+e, row=85, value = Def_EECName[e])
        ws2.cell(column = 2+2*11+e, row=85, value = Def_EECName[e])  
        ws2.cell(column = 2+3*11+e, row=85, value = Def_EECName[e])
        ws2.cell(column = 2+4*11+e, row=85, value = Def_EECName[e])
        
    #Data
    for t in range (24, Par_NoOfYears-1):
     for e in range (0, Par_NoOfEEC):
         
        ws2.cell(column = 2+0*11+e, row=86+t-24, value = Total_Stock[t,e])
        ws2.cell(column = 2+1*11+e, row=86+t-24, value = Total_EoL[t,e])
        ws2.cell(column = 2+2*11+e, row=86+t-24, value = Total_Recycling[t,e])
        ws2.cell(column = 2+3*11+e, row=86+t-24, value = Total_Reuse[t,e])
        ws2.cell(column = 2+4*11+e, row=86+t-24, value = Total_Production[t,e])
        
        ws2.cell(column = 11+0*11, row=86+t-24).border = bd
        ws2.cell(column = 11+1*11, row=86+t-24).border = bd
        ws2.cell(column = 11+2*11, row=86+t-24).border = bd
        ws2.cell(column = 11+3*11, row=86+t-24).border = bd
        ws2.cell(column = 11+4*11, row=86+t-24).border = bd   
     
    ##End Total
        
    
    #%%
    ##Generate new sheets for LCA Data
    
    for e in range (Par_NoOfImpactCat):
        name =  Def_ImpactCatCode[e]
        ws3 = Result_WB.create_sheet(name)
        
        myfont = Font(name='Arial',
                         size=13,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000') 
        
        bd = Border(left=Side(style=None), 
                             right=Side(style='thick'), 
                             top=Side(style=None), 
                             bottom=Side(style=None))
        ws3['A1'] = "Dynamic Stock Model for WEEE"
        ws3['A2'] = name
        ws3['E1'] = "Created: "+ str(datetime.now())
        ws3['E2'] = "Selected Reuse Scenario: Scenario "+str(sc+1)
        
        ws3['A1'].font = myfont
        ws3['A2'].font = myfont
        ws3['E1'].font = myfont
        
        ##Print Impacts
        
        #Years
        for a in range (24, Par_NoOfYears-1): 
                 ws3.cell(column=1, row = a+6-24, value = Def_Time[a])
                 ws3.cell(column=1, row = a+6-24).font = myfont 
                
        #Processes
        ws3.cell(column=2, row = 5, value = "ImpactsTotal")
        ws3.cell(column=2, row = 5).font = myfont    
        
        ws3.cell(column=3, row = 5, value = "ImpactsUse")
        ws3.cell(column=3, row = 5).font = myfont 
        
        ws3.cell(column=4, row = 5, value = "ImpactsTotalNew")
        ws3.cell(column=4, row = 5).font = myfont 
        
        ws3.cell(column=5, row = 5, value = "ImpactsTotalPfR")
        ws3.cell(column=5, row = 5).font = myfont
        
        ws3.cell(column=6, row = 5, value = "ImpactsRecycling")
        ws3.cell(column=6, row = 5).font = myfont 
        
        ws3.cell(column=8, row = 5, value = "ImpactsProdNew")
        ws3.cell(column=8, row = 5).font = myfont 
        
        ws3.cell(column=9, row = 5, value = "ImpactsTransportNew")
        ws3.cell(column=9, row = 5).font = myfont 
         
        ws3.cell(column=10, row = 5, value = "ImpactsRepairPfR")
        ws3.cell(column=10, row = 5).font = myfont 
        
        ws3.cell(column=11, row = 5, value = "ImpactsTransportPfR")
        ws3.cell(column=11, row = 5).font = myfont 
          
        
        for a in range (24, Par_NoOfYears-1):
          ws3.cell(column=2, row = a+6-24, value = Imp_Total [a,e])
          ws3.cell(column=3, row = a+6-24, value = Def_YIU [a,e])
          ws3.cell(column=4, row = a+6-24, value = Imp_Totalnew [a,e])
          ws3.cell(column=5, row = a+6-24, value = Imp_TotalPfR [a,e])
          ws3.cell(column=6, row = a+6-24, value = Imp_Recperyear [a,e])
          
          ws3.cell(column=8, row = a+6-24, value = Imp_Pperyear [a,e])
          ws3.cell(column=9, row = a+6-24, value = Imp_Tnew [a,e])
          ws3.cell(column=10, row = a+6-24, value = Imp_R [a,e])
          ws3.cell(column=11, row = a+6-24, value = Imp_TPfR [a,e])
      
        #End LCA Data
        
    now=datetime.now().strftime("%Y_%m_%d %H.%M.%S")
    
    print (now)

    dest_filename = os.path.join(Path_Run, 'Results DSM_Sc' + str(sc + 1) + '_' + str(now) +'.xlsx')
    Result_WB.save(filename = dest_filename)
    
    print("Workbook saved.")

             

            
             
        
